from wialon.sdk import WialonSdk, WialonError, SdkException
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
import datetime as dt

# Логин по токену:
sdk = WialonSdk(
    is_development=True,
    scheme='https',
    host='maps.baltgps.ru',
    port=0,
    session_id='',
    extra_params={}
)
token = "4c0646a576604f545368e8187177098fBB0550C4B2647584535F2EA745811639516044F9"
response = sdk.login(token)

wb = Workbook()
ws = wb.active
ws['A1'] = "Название объекта"
ws['B1'] = "Пробег за день, км"
time_now = str(dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
now = dt.datetime.strptime(time_now, '%Y-%m-%d %H:%M:%S')
time_to_unix = now.timestamp()

"""1 - Дата сегодня в формате ГГ-ММ-ДД(str); 2 - Начало сегодняшнего дня,00-00(str);
3 - из строки day_at_00 делаю datetime объект; 4 - получаю UNIX - время(float)"""

today_day = dt.datetime.now().strftime('%Y-%m-%d')
day_at_00 = today_day+" 00:00:00"
time_from = dt.datetime.strptime(day_at_00, '%Y-%m-%d %H:%M:%S')
time_from_unix = time_from.timestamp()

send_list = []
qwe = len(send_list)
# Выполняю отчет "Пробег по дням (групповой)"(13) по группе "Дягилев А.А. (гор)"
try:
    params = {"reportResourceId": 16512, "reportTemplateId": 13, "reportObjectId": 20805,
              "reportObjectSecId": 0, "interval": {"from": int(time_from_unix), "to": int(time_to_unix), "flags": 0}}
    report = sdk.report_exec_report(params)

    # Выгружаю данные по отчету
    params = {"tableIndex": 0, "indexFrom": 0, "indexTo": 100}
    report_table = sdk.report_get_result_rows(params)

    # По каждой машине сравниваю пробег, если больше  350 км ,то записываю "Объект" + "Пробег" в отчет
    for car in report_table:
        mileage_per_day = float(car["c"][1].replace(" ","").replace("km",""))
        car_name = car["c"][0]
        if mileage_per_day > 350:
            note = [car_name, mileage_per_day]
            send_list.append(note)
            ws.append(note)

except SdkException as e:
    print(f'Sdk related error: {e}')
except WialonError as e:
    print(f'Wialon related error: {e}')
except Exception as e:
    print(f'Python error: {e}')

# Сохраняю excel-файл:
join = (today_day, ".xlsx")
excel_name = str("".join(join))
wb.save(excel_name)

# Отправка email:
if len(send_list) > 0:
    sender = "**********@*********"
    receiver = "************@**************"
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = receiver
    msg['Subject'] = "Отчет"
    body = "В прикрепленном файле отчет по машинам за сегодня."
    msg.attach(MIMEText(body, 'plain'))
    filename = excel_name
    attachment = open("report_2020-08-13_.xlsx", "rb")
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename=%s" % filename)
    msg.attach(p)
    s = smtplib.SMTP('smtp.gmail.com', 587)
    s.starttls()
    s.login(**************, "**************")
    text = msg.as_string()
    s.sendmail(sender, receiver, text)
else:
    print("Пробег всех машин не превышает 350 км, письмо e-mail не создано.")
